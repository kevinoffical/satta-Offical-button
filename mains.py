import os
import requests
from bs4 import BeautifulSoup
from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import JSONResponse, HTMLResponse
import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import logging
import csv

app = FastAPI()
logging.basicConfig(level=logging.INFO)

# Telegram bot token from environment variable
TOKEN = os.environ.get('TOKEN')
if not TOKEN:
    raise ValueError("Bot token not set in environment variables. Please set the 'TOKEN' variable.")

bot = telebot.TeleBot(TOKEN)

@app.post('/webhook/')
async def webhook(request: Request):
    json_str = await request.json()
    update = telebot.types.Update.de_json(json_str)
    bot.process_new_updates([update])
    return JSONResponse(content={"status": "ok"})

@app.get('/')
async def index():
    return HTMLResponse(content="Bot is Live", status_code=200)

@app.head('/')
async def index_head():
    return HTMLResponse(content="Bot is Live", status_code=200)

# Constants
URL = "https://satta-king-fast.com/"
TIMEZONE = 'Asia/Kolkata'

# Emoji constants
EMOJI_CALENDAR = '📅'
EMOJI_ROBOT = '🤖'
EMOJI_BACK = '↩️'
EMOJI_STOP = '🛑'
EMOJI_SUCCESS = '✅'

# Mapping of game names to display names and emojis
GAME_NAMES = {
    'DSWR': {'name': 'DESAWAR', 'emoji': '🎲'},
    'FRBD': {'name': 'FARIDABAD', 'emoji': '🎲'},
    'GZBD': {'name': 'GHAZIABAD', 'emoji': '🎲'},
    'GALI': {'name': 'GALI', 'emoji': '🎲'},
}

# Dictionary to store user data
user_data = {}

def get_current_time():
    """Get current date and time in IST."""
    ist = pytz.timezone(TIMEZONE)
    utc_now = datetime.utcnow().replace(tzinfo=pytz.utc)
    ist_now = utc_now.astimezone(ist)
    return ist_now

# Start command handler
@bot.message_handler(commands=['start'])
def send_start(message):
    username = message.from_user.first_name
    ist_now = get_current_time()
    ist_formatted = ist_now.strftime('%d %B %Y %I:%M:%S %p')

    welcome_message = (
        f"👋 Hello *{username}*! 🤖\n\n"
        f"Welcome to the *Satta King Bot*! ✅\n\n"
        f"📅 Current Date & Time: \n*{ist_formatted}*\n\n"
        f"Use the buttons below to get started:"
    )

    markup = InlineKeyboardMarkup(row_width=2)
    buttons = [
        InlineKeyboardButton("Get Chart 📊", callback_data='chart'),
        InlineKeyboardButton("Get Prediction 🔮", callback_data='predict'),
        InlineKeyboardButton("Check My Number 🔍", callback_data='checkmynumber'),
        InlineKeyboardButton("Close 🛑", callback_data='close')
    ]
    markup.add(*buttons)

    sent_message = bot.send_message(message.chat.id, welcome_message, parse_mode='Markdown', reply_markup=markup)
    user_data[message.chat.id] = {"message_id": sent_message.message_id}

# Update message with new content and markup
def update_message(chat_id, message_id, new_text, new_markup):
    bot.edit_message_text(new_text, chat_id, message_id, reply_markup=new_markup, parse_mode='Markdown')

# Chart button handler
def handle_chart(message):
    ist_now = get_current_time().strftime('%Y-%m-%d %H:%M:%S')

    chart_message = (
        f"*{EMOJI_ROBOT} Welcome to the Satta King Chart Section.*\n\n"
        f"{EMOJI_CALENDAR} Current Date & Time: \n*{ist_now}*\n\n"
        f"Please select the year for\nwhich you want the chart data:"
    )

    markup = InlineKeyboardMarkup(row_width=3)
    years = [str(year) for year in range(2015, 2025)]
    year_buttons = [InlineKeyboardButton(year, callback_data=f"year_{year}") for year in years]
    back_button = InlineKeyboardButton(f"{EMOJI_BACK} Back", callback_data='back_to_start')
    close_button = InlineKeyboardButton(f"{EMOJI_STOP} Close", callback_data='close')
    markup.add(*year_buttons)
    markup.add(back_button, close_button)

    update_message(message.chat.id, user_data[message.chat.id]["message_id"], chart_message, markup)

# Predict button handler
def handle_predict(message):
    ist_now = get_current_time()
    formatted_date = ist_now.strftime('%d %B %Y')
    formatted_time = ist_now.strftime('%I:%M:%S %p')

    predict_message = (
        f"{EMOJI_ROBOT} Welcome to the Satta King Prediction Section.\n\n"
        f"{EMOJI_CALENDAR} Current Date: {formatted_date}\n"
        f"{EMOJI_CALENDAR} Current Time (IST): {formatted_time}\n\n"
        "Please select which game's prediction you want:"
    )

    markup = InlineKeyboardMarkup(row_width=2)
    game_buttons = [
        InlineKeyboardButton(game_info['name'], callback_data=f'predict_{code}')
        for code, game_info in GAME_NAMES.items()
    ]
    back_button = InlineKeyboardButton(f"{EMOJI_BACK} Back", callback_data='back_to_start')
    close_button = InlineKeyboardButton(f"{EMOJI_STOP} Close", callback_data='close')
    markup.add(*game_buttons)
    markup.add(back_button, close_button)

    update_message(message.chat.id, user_data[message.chat.id]["message_id"], predict_message, markup)

# Check My Number button handler
def handle_checkmynumber(message):
    number_prompt = "Tell me your number (between 00 and 99):"

    markup = InlineKeyboardMarkup()
    back_button = InlineKeyboardButton(f"{EMOJI_BACK} Back", callback_data='back_to_start')
    close_button = InlineKeyboardButton(f"{EMOJI_STOP} Close", callback_data='close')
    markup.add(back_button, close_button)

    sent_message = bot.edit_message_text(number_prompt, message.chat.id, user_data[message.chat.id]["message_id"], reply_markup=markup)
    bot.register_next_step_handler(sent_message, get_user_number)

def get_user_number(message):
    try:
        user_number = message.text.strip()
        if user_number.isdigit() and 0 <= int(user_number) <= 99:
            user_data[message.chat.id] = {"number": user_number}

            markup = InlineKeyboardMarkup(row_width=3)
            months_buttons = [
                InlineKeyboardButton(f"{months} months", callback_data=f"number_months_{months}")
                for months in range(6, 121, 6)
            ]
            close_button = InlineKeyboardButton(f"{EMOJI_STOP} Close", callback_data='close')
            markup.add(*months_buttons)
            markup.add(close_button)

            bot.send_message(message.chat.id, "Select range for chart detail:", reply_markup=markup)
        else:
            skip_message = "Invalid number provided. Skipping this task. You can try again using the buttons below."
            bot.send_message(message.chat.id, skip_message)
            send_start(message)  # Re-initiate the start command
    except Exception as e:
        logging.error(f"Error: {str(e)}")
        skip_message = "An error occurred. Skipping this task."
        bot.send_message(message.chat.id, skip_message)
        send_start(message)  # Re-initiate the start command

# Callback query handler
@bot.callback_query_handler(func=lambda call: True)
def handle_callback(call):
    try:
        if call.data == 'chart':
            handle_chart(call.message)
        elif call.data == 'predict':
            handle_predict(call.message)
        elif call.data == 'checkmynumber':
            handle_checkmynumber(call.message)
        elif call.data == 'close':
            bot.delete_message(call.message.chat.id, call.message.message_id)
        elif call.data == 'back_to_start':
            bot.delete_message(call.message.chat.id, call.message.message_id)
            send_start(call.message)
        elif call.data.startswith('year_'):
            year = call.data.split('_')[1]
            show_month_selection(call.message, year)
        elif call.data.startswith('month_'):
            process_month_selection(call)
        elif call.data.startswith('predict_'):
            handle_prediction_query(call)
        elif call.data == 'show_latest_number':
            show_latest_number(call)
        elif call.data.startswith('months_'):
            handle_months_selection(call)
        elif call.data.startswith('number_months_'):
            handle_number_months_selection(call)
        elif call.data == 'back_to_year_selection':
            handle_chart(call.message)
    except Exception as e:
        logging.error(f"Callback error: {str(e)}")

def show_month_selection(message, year):
    markup = InlineKeyboardMarkup(row_width=3)
    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    buttons = [InlineKeyboardButton(month, callback_data=f"month_{month.lower()}_{year}") for month in months]
    back_button = InlineKeyboardButton(f"{EMOJI_BACK} Back to Year Selection", callback_data="back_to_year_selection")
    close_button = InlineKeyboardButton(f"{EMOJI_STOP} Close", callback_data='close')
    markup.add(*buttons)
    markup.add(back_button, close_button)
    bot.edit_message_text(f"Select the month for {year}:", message.chat.id, user_data[message.chat.id]["message_id"], reply_markup=markup)

def process_month_selection(call):
    month, year = call.data.split('_')[1:]
    month_number = {
        "january": "01", "february": "02", "march": "03", "april": "04", "may": "05", "june": "06",
        "july": "07", "august": "08", "september": "09", "october": "10", "november": "11", "december": "12"
    }[month]

    url = f"https://satta-king-fast.com/chart.php?month={month_number}&year={year}"

    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        table = soup.find('table', class_='chart-table')
        if not table:
            bot.send_message(call.message.chat.id, f"No data found for {month.capitalize()} {year}")
            return

        rows = table.find_all('tr')
        if not rows or len(rows) < 3:
            bot.send_message(call.message.chat.id, f"No data available in the table for {month.capitalize()} {year}")
            return

        header = [cell.text.strip() for cell in rows[1].find_all(['th', 'td'])]
        data = []

        for row in rows[2:]:
            cells = row.find_all(['th', 'td'])
            if len(cells) != len(header):
                continue
            date = cells[0].text.strip()
            values = [cell.text.strip() for cell in cells[1:]]
            data.append([date] + values)

        if not data:
            bot.send_message(call.message.chat.id, f"No data rows found for {month.capitalize()} {year}")
            return

        filename = f"Satta_King_Chart_{month.capitalize()}_{year}.csv"
        with open(filename, mode='w', newline='') as file:
            writer = csv.writer(file, delimiter=',')
            writer.writerow(header)
            writer.writerows(data)

        with open(filename, 'r') as file:
            csv_data = file.read()
            formatted_data = format_chart_data(csv_data)
            chart_message = f"Here is the Satta King Chart for {month.capitalize()} {year}:\n\n{formatted_data}"
            back_button = InlineKeyboardButton(f"{EMOJI_BACK} Back to Year Selection", callback_data="back_to_year_selection")
            close_button = InlineKeyboardButton(f"{EMOJI_STOP} Close", callback_data='close')
            markup = InlineKeyboardMarkup().add(back_button, close_button)
            bot.edit_message_text(chart_message, call.message.chat.id, user_data[call.message.chat.id]["message_id"], reply_markup=markup)

        with open(filename, 'rb') as file:
            bot.send_document(call.message.chat.id, file)

        os.remove(filename)
    except requests.exceptions.RequestException as e:
        bot.send_message(call.message.chat.id, f"Error: {str(e)}")
    except Exception as e:
        bot.send_message(call.message.chat.id, f"Error: {str(e)}")

def format_chart_data(csv_data):
    lines = csv_data.strip().split('\n')
    formatted_lines = []
    for line in lines:
        formatted_line = '    '.join(line.split(','))
        formatted_lines.append(formatted_line)
    return '\n'.join(formatted_lines)

def handle_prediction_query(call):
    game_code = call.data.split('_')[1]
    game_info = GAME_NAMES.get(game_code)

    try:
        response = requests.get(URL)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, 'html.parser')
        game_element = soup.find('h3', class_='game-name', string=game_info['name'])
        if not game_element:
            raise ValueError(f"Game {game_info['name']} not found on the website.")

        today_number_element = game_element.find_next('td', class_='today-number').find('h3')
        today_number = today_number_element.text.strip() if today_number_element else None

        yesterday_number_element = game_element.find_next('td', class_='yesterday-number').find('h3')
        yesterday_number = yesterday_number_element.text.strip() if yesterday_number_element else None

        yesterday_time_element = game_element.find_next('h3', class_='game-time').text.strip() if game_element.find_next('h3', class_='game-time') else None

        ist_now = get_current_time()
        formatted_date_today = ist_now.strftime('%d %B %Y')
        formatted_time_today = ist_now.strftime('%I:%M:%S %p')

        ist_yesterday = ist_now - timedelta(days=1)
        formatted_date_yesterday = ist_yesterday.strftime('%d %B %Y')

        if today_number and today_number != 'XX':
            prediction_message = (
                f"Today's number for {game_info['name']} ({game_code}) is:\n"
                f"{today_number} {game_info['emoji']}\n\n"
                f"{EMOJI_CALENDAR} Today's Date: {formatted_date_today}\n"
                f"{EMOJI_CALENDAR} Today's Time (IST): {formatted_time_today}\n\n"
                f"{EMOJI_CALENDAR} Yesterday's Date: {formatted_date_yesterday}\n"
                f"{EMOJI_CALENDAR} Yesterday's Time (IST): {yesterday_time_element}"
            )
            latest_number = today_number
        else:
            prediction_message = (
                f"Yesterday's number for {game_info['name']} ({game_code}) was:\n"
                f"{yesterday_number} {game_info['emoji']}\n\n"
                f"{EMOJI_CALENDAR} Yesterday's Date: {formatted_date_yesterday}\n"
                f"{EMOJI_CALENDAR} Yesterday's Time (IST): {yesterday_time_element}\n\n"
                f"{EMOJI_CALENDAR} Today's Date: {formatted_date_today}\n"
                f"{EMOJI_CALENDAR} Today's Time (IST): {formatted_time_today}"
            )
            latest_number = yesterday_number

        markup = InlineKeyboardMarkup(row_width=2)
        back_button = InlineKeyboardButton(f"{EMOJI_BACK} Back", callback_data="back_to_start")
        latest_number_button = InlineKeyboardButton(f"Check Chart {latest_number} 🎲", callback_data="show_latest_number")
        close_button = InlineKeyboardButton(f"{EMOJI_STOP} Close", callback_data="close")

        markup.add(latest_number_button, back_button)
        markup.add(close_button)

        user_data[call.message.chat.id]['latest_number'] = latest_number
        bot.edit_message_text(prediction_message, call.message.chat.id, call.message.message_id, reply_markup=markup)

    except requests.exceptions.RequestException as e:
        error_message = f"Error fetching data from the website: {str(e)}"
        logging.error(error_message)
        bot.send_message(call.message.chat.id, error_message)
    except Exception as e:
        error_message = f"Error fetching prediction for {game_info['name']} ({game_code}): {str(e)}"
        logging.error(error_message)
        bot.send_message(call.message.chat.id, error_message)

def show_latest_number(call):
    try:
        markup = InlineKeyboardMarkup(row_width=3)
        months_buttons = [
            InlineKeyboardButton(f"{months} months", callback_data=f"months_{months}")
            for months in range(6, 121, 6)
        ]
        markup.add(*months_buttons)
        markup.add(InlineKeyboardButton(f"{EMOJI_BACK} Back", callback_data="back_to_start"))

        bot.edit_message_text("Select the range for chart data:", call.message.chat.id, call.message.message_id, reply_markup=markup)

    except Exception as e:
        error_message = f"Error: {str(e)}"
        logging.error(error_message)
        bot.send_message(call.message.chat.id, error_message)

def handle_months_selection(call):
    try:
        months = int(call.data.split('_')[1])
        preparing_message = bot.send_message(call.message.chat.id, "Please wait, preparing your file...")
        file_path = fetch_chart_data_for_months(months, user_data[call.message.chat.id])
        bot.delete_message(call.message.chat.id, preparing_message.message_id)
        bot.send_document(call.message.chat.id, open(file_path, 'rb'))
        os.remove(file_path)
    except Exception as e:
        error_message = f"Error: {str(e)}"
        logging.error(error_message)
        bot.send_message(call.message.chat.id, error_message)

def fetch_chart_data_for_months(months, user_data):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Satta King Chart Last {months} Months"

        headers = ['DATE'] + [game_info['name'] for game_info in GAME_NAMES.values()]
        header_font = Font(size=12, bold=True)
        header_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        ws.append(headers)
        for col_num, cell in enumerate(ws[1], start=1):
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[chr(64 + col_num)].width = 15

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        current_date = datetime.now()
        latest_number = user_data.get('latest_number')

        for i in range(months):
            month = (current_date - timedelta(days=i*30)).month
            year = (current_date - timedelta(days=i*30)).year

            chart_url = f"https://satta-king-fast.com/chart.php?month={month:02}&year={year}"
            response = requests.get(chart_url)
            response.raise_for_status()

            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', class_='chart-table')
            if not table:
                raise ValueError(f"No data table found for {month}-{year}")

            rows = table.find_all('tr')
            if len(rows) < 3:
                raise ValueError(f"Insufficient data rows in the table for {month}-{year}")

            ws.append([f"{datetime(year, month, 1).strftime('%B-%Y')}"])

            for row in rows[2:]:
                date_cell = row.find('td', class_='day')
                if not date_cell:
                    continue

                date = date_cell.text.strip()
                numbers = [cell.text.strip() for cell in row.find_all('td', class_='number')]

                row_data = [date] + numbers
                ws.append(row_data)

                if latest_number and latest_number in numbers:
                    col_num = numbers.index(latest_number) + 2
                    ws.cell(row=ws.max_row, column=col_num).fill = yellow_fill

                ws.row_dimensions[ws.max_row].height = 17

        file_path = f"satta_king_last_{months}_months.xlsx"
        wb.save(file_path)

        return file_path

    except requests.exceptions.RequestException as e:
        error_message = f"Error fetching data from the website: {str(e)}"
        logging.error(error_message)
        raise
    except Exception as e:
        error_message = f"Error generating chart: {str(e)}"
        logging.error(error_message)
        raise

def handle_number_months_selection(call):
    try:
        months = int(call.data.split('_')[2])
        user_number = user_data[call.message.chat.id]['number']
        preparing_message = bot.send_message(call.message.chat.id, "Please wait, preparing your file...")
        file_path = fetch_chart_data_for_months(months, {"latest_number": user_number})
        bot.delete_message(call.message.chat.id, preparing_message.message_id)
        bot.send_document(call.message.chat.id, open(file_path, 'rb'))
        os.remove(file_path)
    except Exception as e:
        error_message = f"Error: {str(e)}"
        logging.error(error_message)
        bot.send_message(call.message.chat.id, error_message)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
